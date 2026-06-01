"""
Skill 仓库同步脚本 v2.0
=========================
功能：
  1. 自动扫描 ~/.agents/skills/ 和 ~/.claude/skills/ 下所有已安装的 Skill
  2. 智能分类（基于 skill-classifier.json 规则 + SKILL.md 内容分析）
  3. 同步更新到 Skill仓库.xlsx（精美格式）
  4. 生成 skills-index.json（供推荐引擎快速查询）

用法：
  python scripts/sync-skills-repo.py           # 同步
  python scripts/sync-skills-repo.py --watch   # 持续监听（未实现）
"""
import os
import sys
import json
import re
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 路径配置 ────────────────────────────────────────────────
CWD = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CWD)
CLASSIFIER_PATH = os.path.join(CWD, 'skill-classifier.json')
INDEX_PATH = os.path.join(PROJECT_ROOT, 'skills-index.json')
OUTPUT = os.path.join(PROJECT_ROOT, 'Skill仓库.xlsx')

SKILLS_DIRS = [
    os.path.join(os.path.expanduser('~'), '.agents', 'skills'),
    os.path.join(os.path.expanduser('~'), '.claude', 'skills'),
]

# ── 颜色配置 ────────────────────────────────────────────────
CATEGORY_COLORS = {
    '前端开发':        ('1B8132', 'C6EFCE'),
    '后端开发':        ('2F5496', 'D6E4F0'),
    '飞书/Lark办公':   ('BF8F00', 'FFF2CC'),
    'AI/机器学习':     ('7030A0', 'E2D0F0'),
    'DevOps/部署':     ('C00000', 'F4CCCC'),
    '数据库':          ('0070C0', 'DDEBF7'),
    '测试':            ('548235', 'E2EFDA'),
    '文档写作':        ('808080', 'F2F2F2'),
    '设计/UI/UX':     ('E36C0A', 'FCE4D6'),
    '代码质量/安全':   ('C65911', 'FCE4D6'),
    '项目管理':        ('1F4E79', 'D6E4F0'),
    '移动开发':        ('0B8043', 'D9EAD3'),
    '性能优化':        ('6E6E6E', 'EFEFEF'),
    '数据分析':        ('38761D', 'D9EAD3'),
    '技能管理':        ('4A86E8', 'C9DAF8'),
    '办公自动化':      ('783F04', 'FCE5CD'),
    '其他':            ('999999', 'F7F7F7'),
}


def load_classifier():
    """加载分类规则"""
    if os.path.exists(CLASSIFIER_PATH):
        with open(CLASSIFIER_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def classify_skill(name, desc, dirname, classifier):
    """
    智能分类：基于多维度关键词匹配 + 加权评分
    返回 (category, confidence)
    """
    if classifier is None:
        return fallback_classify(dirname), 50

    scores = {}
    cat_config = classifier['categories']

    # 统一转为小写进行匹配
    name_lower = name.lower()
    desc_lower = desc.lower()
    dir_lower = dirname.lower()

    for cat_name, cat_rules in cat_config.items():
        score = 0
        weight = cat_rules['weight']
        kw = cat_rules['keywords']

        # 名称匹配（权重最高）
        for k in kw.get('name', []):
            if k in name_lower:
                score += 3 * weight

        # 描述匹配
        for k in kw.get('desc', []):
            if k in desc_lower:
                score += 2 * weight

        # 目录名匹配
        for k in kw.get('dir', []):
            if k in dir_lower:
                score += 2 * weight

        if score > 0:
            # 修正：如果描述中有明显的排除词
            if name_lower.startswith('lark-') and cat_name != '飞书/Lark办公':
                score *= 0.3  # lark 技能强烈倾向飞书分类
            if 'skill' in dir_lower and cat_name != '技能管理':
                score *= 0.5

            scores[cat_name] = score

    if not scores:
        return classifier['fallback_category'], 0

    # 取最高分
    best_cat = max(scores, key=scores.get)
    max_score = scores[best_cat]
    # 计算置信度（0-100）
    total_possible = 5 * weight  # 假设单项满分
    second_best = sorted(scores.values(), reverse=True)[0] if len(scores) > 1 else 0
    confidence = min(100, int((max_score / max(1, cat_config[best_cat]['weight'] * 5)) * 70 +
                              (1 - second_best / max(1, max_score)) * 30))

    return best_cat, confidence


def fallback_classify(dirname):
    """兜底分类：基于目录名启发式规则"""
    d = dirname.lower()
    if d.startswith('lark'):
        return '飞书/Lark办公'
    if 'skill' in d:
        return '技能管理'
    if 'front' in d or 'ui' in d or 'design' in d:
        return '前端开发'
    if any(x in d for x in ('test', 'e2e', 'jest')):
        return '测试'
    return '其他'


def extract_metadata(md_path):
    """
    解析 SKILL.md 的 YAML frontmatter
    支持:
    - name: xxx
    - description: xxx
    - description: >
        multi-line folded
    - description: "quoted string"
    """
    name = ''
    desc = ''
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        parts = content.split('---')
        if len(parts) < 3:
            return name, desc

        front_lines = parts[1].split('\n')
        in_desc_block = False
        desc_parts = []

        for line in front_lines:
            s = line.strip()

            if s.startswith('name:'):
                name = s[5:].strip().strip("'").strip('"')
                continue

            if s.startswith('description:'):
                in_desc_block = True
                rest = s[12:].strip()
                # 处理 "> " folded block scalar
                if rest == '>':
                    desc_parts = []
                    continue
                # 处理单行: "xxx" 或 'xxx' 或 xxx
                if rest:
                    desc = rest.strip("'").strip('"').strip()
                continue

            # 处理多行 folded block (description: > 下面的缩进行)
            if in_desc_block and s and not s.startswith('---'):
                if s.startswith((' ', '\t')) or any(
                    c in s for c in 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                ):
                    # 检查是否是下一字段的开始
                    if s.startswith(('name:', 'compatibility:', 'danger:')):
                        in_desc_block = False
                    else:
                        # 移除引号
                        clean = s.strip("'").strip('"').strip()
                        if clean:
                            desc_parts.append(clean)
                        continue
                else:
                    in_desc_block = False

        if desc_parts:
            desc = ' '.join(desc_parts)
        # 清理：合并多余空白
        desc = ' '.join(desc.split())
    except Exception:
        pass
    return name, desc


def get_skill_tags(name, desc, dirname):
    """从技能信息中提取标签（用于筛选和推荐）"""
    text = f"{name} {desc} {dirname}".lower()
    tags = set()

    # 技术栈标签
    tech_map = {
        'react': 'React', 'vue': 'Vue', 'angular': 'Angular',
        'node': 'Node.js', 'python': 'Python', 'java': 'Java',
        'go': 'Go', 'rust': 'Rust', 'typescript': 'TypeScript',
        'docker': 'Docker', 'kubernetes': 'K8s', 'aws': 'AWS',
        'redis': 'Redis', 'postgres': 'PostgreSQL', 'mysql': 'MySQL',
        'mongodb': 'MongoDB', 'flask': 'Flask', 'django': 'Django',
        'fastapi': 'FastAPI', 'nextjs': 'Next.js', 'nuxt': 'Nuxt',
        'tailwind': 'Tailwind', 'lark': '飞书', 'ai': 'AI',
        'llm': 'LLM', 'test': '测试', 'doc': '文档',
        'api': 'API', 'graphql': 'GraphQL', 'ci': 'CI/CD',
        'web': 'Web', 'mobile': '移动端', 'design': '设计',
        'data': '数据', 'perf': '性能', 'security': '安全',
    }
    for keyword, tag in tech_map.items():
        if keyword in text:
            tags.add(tag)

    return sorted(tags)


def collect_skills(classifier):
    """收集所有已安装的技能，并分类"""
    seen = set()
    data = []

    for skills_dir in SKILLS_DIRS:
        if not os.path.isdir(skills_dir):
            continue
        for d in sorted(os.listdir(skills_dir)):
            sp = os.path.join(skills_dir, d)
            if not os.path.isdir(sp) or d.startswith('.'):
                continue
            if d in seen:
                continue
            seen.add(d)

            md = os.path.join(sp, 'SKILL.md')
            name, desc = extract_metadata(md)
            if not name:
                name = d
            if not desc:
                # 尝试从 SKILL.md 正文中提取第一段描述
                desc = extract_fallback_desc(md)

            # 智能分类
            category, confidence = classify_skill(name, desc, d, classifier)
            tags = get_skill_tags(name, desc, d)

            # 获取分类描述
            cat_desc = ''
            if classifier and category in classifier['categories']:
                cat_desc = classifier['categories'][category]['description']

            data.append({
                'name': name,
                'description': desc,
                'dir': d,
                'category': category,
                'confidence': confidence,
                'tags': tags,
                'cat_description': cat_desc,
            })

    return data


def extract_fallback_desc(md_path):
    """从 SKILL.md 正文中提取第一段有意义的描述"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        # 去掉 YAML frontmatter
        parts = content.split('---')
        body = parts[-1] if len(parts) >= 2 else content
        # 找到第一个非标题段落
        lines = body.split('\n')
        for line in lines:
            line = line.strip()
            if line and not line.startswith('#') and not line.startswith('!'):
                return line[:200]
    except Exception:
        pass
    return ''


def build_workbook(data):
    """构建精美格式的 Excel 工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Skill仓库'

    # ── 样式定义 ──
    hdr_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Microsoft YaHei', bold=True, size=16, color='1F3864')
    subtitle_font = Font(name='Microsoft YaHei', size=9, color='808080', italic=True)

    cell_font = Font(name='Microsoft YaHei', size=10)
    name_font = Font(name='Microsoft YaHei', size=10, bold=True)
    dir_font = Font(name='Consolas', size=9, color='666666')
    cell_align = Alignment(vertical='center', wrap_text=True)

    tag_font = Font(name='Microsoft YaHei', size=9, color='2F5496')
    tag_fill = PatternFill(start_color='E8EEF7', end_color='E8EEF7', fill_type='solid')

    bdr = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    thin_bdr = Border(
        left=Side(style='hair', color='D9D9D9'),
        right=Side(style='hair', color='D9D9D9'),
        top=Side(style='hair', color='D9D9D9'),
        bottom=Side(style='hair', color='D9D9D9'),
    )

    # ── Title ──
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = 'Skill 仓库 -- Claude Code 技能总览 v2.0'
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 50

    # ── Subtitle ──
    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f'共 {len(data)} 个技能  |  安装新技能后运行 sync-skills-repo.py 自动同步  |  分类基于 AI 智能分析'
    c.font = subtitle_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── Headers ──
    headers = ['技能名称', '功能描述', '目录名', '分类', '标签', '推荐场景']
    hr = 3
    col_widths = [22, 60, 32, 16, 18, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = bdr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[hr].height = 30

    # ── 按分类排序 ──
    cat_order = ['前端开发', '后端开发', 'AI/机器学习', '飞书/Lark办公', 'DevOps/部署',
                 '数据库', '测试', '代码质量/安全', '设计/UI/UX', '文档写作',
                 '项目管理', '移动开发', '性能优化', '数据分析', '技能管理',
                 '办公自动化', '其他']
    data_sorted = sorted(data, key=lambda x: (
        cat_order.index(x['category']) if x['category'] in cat_order else 99,
        x['name']
    ))

    row = hr + 1
    current_cat = None

    for item in data_sorted:
        cat = item['category']

        # 分类分隔行
        if cat != current_cat:
            current_cat = cat
            hcolor, _ = CATEGORY_COLORS.get(cat, ('999999', 'F7F7F7'))
            cat_fill = PatternFill(start_color=hcolor, end_color=hcolor, fill_type='solid')
            cat_font_styled = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f'▸ {cat}')
            cell.font = cat_font_styled
            cell.fill = cat_fill
            cell.alignment = Alignment(vertical='center')
            cell.border = bdr
            for cc in range(2, 7):
                ws.cell(row=row, column=cc).fill = cat_fill
                ws.cell(row=row, column=cc).border = bdr

            # 分类说明
            if item.get('cat_description'):
                ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
                cd = ws.cell(row=row + 1, column=1,
                             value=f'   -> {item["cat_description"]}')
                cd.font = Font(name='Microsoft YaHei', size=9, color='666666', italic=True)
                cd.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                cd.alignment = Alignment(vertical='center')
                cd.border = thin_bdr
                for cc in range(2, 7):
                    ws.cell(row=row + 1, column=cc).fill = PatternFill(
                        start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                    ws.cell(row=row + 1, column=cc).border = thin_bdr
                ws.row_dimensions[row + 1].height = 20
                row += 1

            ws.row_dimensions[row].height = 28
            row += 1

        # 技能行
        _, bgcolor = CATEGORY_COLORS.get(cat, ('999999', 'F7F7F7'))
        alt_fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type='solid')

        ws.cell(row=row, column=1, value=item['name']).font = name_font
        ws.cell(row=row, column=2, value=item['description']).font = cell_font
        ws.cell(row=row, column=3, value=item['dir']).font = dir_font
        ws.cell(row=row, column=4, value=cat).font = cell_font

        # 标签（用不同底色显示）
        tags_str = ', '.join(item['tags']) if item['tags'] else ''
        ws.cell(row=row, column=5, value=tags_str).font = tag_font

        # 推荐场景（基于技能名称和描述生成简短推荐场景）
        scene = generate_scene(item)
        ws.cell(row=row, column=6, value=scene).font = Font(name='Microsoft YaHei', size=9, color='444444')

        for cc in range(1, 7):
            ws.cell(row=row, column=cc).alignment = cell_align
            ws.cell(row=row, column=cc).border = thin_bdr
            if cc != 1 and cc != 6:
                ws.cell(row=row, column=cc).fill = alt_fill

        # 行高自适应
        desc_len = len(item['description'])
        ws.row_dimensions[row].height = max(28, 14 * (desc_len // 55 + 1))
        row += 1

    # ── 冻结窗格 ──
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:F{row - 1}'

    # ── 图例行 ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    legend = ws.cell(row=row, column=1,
                     value='安装新技能后自动同步 (PostToolUse hook) | 手动同步: python scripts/sync-skills-repo.py')
    legend.font = Font(name='Microsoft YaHei', size=9, color='999999', italic=True)
    legend.alignment = Alignment(horizontal='center')

    return wb


def generate_scene(item):
    """生成简短推荐场景描述"""
    name = item['name'].lower()
    desc = item['description'].lower()
    cat = item['category']

    if 'lark' in name:
        return '飞书办公场景'
    if cat == '技能管理':
        return '技能发现/创建/管理'
    if 'test' in name or '测试' in desc:
        return '代码测试/质量保障'
    if 'doc' in name or '文档' in desc:
        return '文档编写/维护'
    if 'deploy' in name or '部署' in desc:
        return '部署运维场景'
    if 'design' in name or '设计' in desc:
        return '界面设计场景'
    if 'review' in name:
        return '代码审查场景'
    if 'ai' in name or 'llm' in name or '大模型' in desc:
        return 'AI 开发场景'
    if 'data' in name or '数据' in desc:
        return '数据分析场景'

    return '通用开发场景'


def build_index(data):
    """构建 JSON 索引文件（供推荐引擎快速查询）"""
    index = {
        'version': '2.0',
        'total': len(data),
        'skills': data,
        'categories': {},
    }

    # 按分类索引
    for item in data:
        cat = item['category']
        if cat not in index['categories']:
            index['categories'][cat] = []
        index['categories'][cat].append(item['name'])

    return index


def main():
    is_first_run = not os.path.exists(OUTPUT)

    print('=' * 60)
    print('  Skill 仓库同步 v2.0')
    print('=' * 60)

    if is_first_run:
        print()
        print('  [*] 检测到首次运行，正在自动创建 Skill 仓库...')
        print('  [*] 将扫描本机所有已安装的 AI 技能并分类入库')
        print()

    # 加载分类器
    classifier = load_classifier()
    if classifier:
        print(f'  [OK] 加载分类规则: {len(classifier["categories"])} 个分类')
    else:
        print('  [!] 未找到分类规则文件，使用兜底分类')
        classifier = None

    # 收集技能
    data = collect_skills(classifier)
    if not data:
        print('  [X] 未找到任何技能')
        print('  [!] 请先安装一些技能: npx skills add <package>')
        sys.exit(1)

    print(f'  [OK] 找到 {len(data)} 个技能')

    # 统计分类
    cat_counts = {}
    for item in data:
        cat = item['category']
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    print(f'  [OK] 分类统计:')
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f'      {cat}: {count}')

    # 生成 Excel
    wb = build_workbook(data)
    wb.save(OUTPUT)
    print(f'  [OK] Excel 已保存: {OUTPUT}')

    # 生成 JSON 索引
    index = build_index(data)
    with open(INDEX_PATH, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    print(f'  [OK] JSON 索引已保存: {INDEX_PATH}')

    print('=' * 60)

    if is_first_run:
        print('  Skill 仓库创建成功!')
        print()
        print('  后续使用:')
        print('    1. 安装新技能后 -> 自动同步 (已配置 PostToolUse hook)')
        print('    2. 手动同步 -> python scripts/sync-skills-repo.py')
        print('    3. 推荐技能 -> python scripts/skill-recommend.py --project .')
    else:
        print('  同步完成!')

    print('=' * 60)


if __name__ == '__main__':
    main()
